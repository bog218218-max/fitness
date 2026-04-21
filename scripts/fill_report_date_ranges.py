#!/Users/bogdanromanenko/Desktop/тренировки/.venv/bin/python
from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path("/Users/bogdanromanenko/Desktop/тренировки/программа.xlsx")


def build_ranges() -> dict[int, str]:
    start = date(2026, 3, 30)
    ranges: dict[int, str] = {}
    for month_index in range(1, 7):
        range_start = start + timedelta(days=28 * (month_index - 1))
        range_end = range_start + timedelta(days=28)
        ranges[month_index] = f"{range_start.strftime('%d.%m.%Y')} - {range_end.strftime('%d.%m.%Y')}"
    return ranges


def main() -> None:
    ranges = build_ranges()
    wb = load_workbook(WORKBOOK_PATH)

    for sheet_name in ["Вторник — Отчет", "Пятница — Отчет"]:
        ws = wb[sheet_name]
        for row in range(1, ws.max_row + 1):
            month_value = ws.cell(row, 2).value
            if isinstance(month_value, (int, float)):
                month_number = int(month_value)
                if month_number in ranges:
                    ws.cell(row, 1).value = ranges[month_number]

    wb.save(WORKBOOK_PATH)
    wb.close()


if __name__ == "__main__":
    main()
