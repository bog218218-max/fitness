#!/Users/bogdanromanenko/Desktop/тренировки/.venv/bin/python
from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path("/Users/bogdanromanenko/Desktop/тренировки/программа.xlsx")


def find_title_row(ws, title_prefix: str) -> int:
    for row in range(1, ws.max_row + 1):
        value = ws[f"A{row}"].value
        if isinstance(value, str) and value.startswith(title_prefix):
            return row
    raise ValueError(f"Title not found: {title_prefix}")


def fill_entry(ws, title_prefix: str, date: str, week: int, sets: list[tuple[float | int | None, int | None, int | None]], summary: str = "", notes: str = "") -> None:
    title_row = find_title_row(ws, title_prefix)
    row = title_row + 2

    ws.cell(row, 1, date)
    ws.cell(row, 2, week)

    col = 3
    for weight, reps, rir in sets[:4]:
        if weight is not None:
            ws.cell(row, col, weight)
        if reps is not None:
            ws.cell(row, col + 1, reps)
        if rir is not None:
            ws.cell(row, col + 2, rir)
        col += 3

    if summary:
        ws.cell(row, 15, summary)
    if notes:
        ws.cell(row, 16, notes)


def append_exercise_block(ws, title: str, target: str) -> None:
    title_row = ws.max_row + 2
    header_row = title_row + 1
    data_row = title_row + 2

    # Reuse styles from the first existing exercise block on the sheet.
    source_title_row = 5
    source_header_row = 6
    source_data_row = 7

    ws.cell(title_row, 1, f"{title} | {target}")
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=16)
    ws.row_dimensions[title_row].height = ws.row_dimensions[source_title_row].height
    ws["A1"]  # touch sheet so dimensions update

    for col in range(1, 17):
        source = ws.cell(source_header_row, col)
        target_cell = ws.cell(header_row, col, source.value)
        if source.has_style:
            target_cell._style = copy(source._style)
        if source.font:
            target_cell.font = copy(source.font)
        if source.fill:
            target_cell.fill = copy(source.fill)
        if source.alignment:
            target_cell.alignment = copy(source.alignment)

        data_source = ws.cell(source_data_row, col)
        data_cell = ws.cell(data_row, col, "")
        if data_source.has_style:
            data_cell._style = copy(data_source._style)
        if data_source.font:
            data_cell.font = copy(data_source.font)
        if data_source.fill:
            data_cell.fill = copy(data_source.fill)
        if data_source.alignment:
            data_cell.alignment = copy(data_source.alignment)

    title_source = ws.cell(source_title_row, 1)
    title_cell = ws.cell(title_row, 1)
    if title_source.has_style:
        title_cell._style = copy(title_source._style)
    if title_source.font:
        title_cell.font = copy(title_source.font)
    if title_source.fill:
        title_cell.fill = copy(title_source.fill)
    if title_source.alignment:
        title_cell.alignment = copy(title_source.alignment)

    ws.row_dimensions[header_row].height = ws.row_dimensions[source_header_row].height
    ws.row_dimensions[data_row].height = ws.row_dimensions[source_data_row].height


def ensure_extra_press_block(ws) -> None:
    try:
        find_title_row(ws, "Пресс (дополнительно)")
    except ValueError:
        append_exercise_block(ws, "Пресс (дополнительно)", "фактический доп. блок")


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH)

    day1 = wb["Вторник — День 1"]
    day2 = wb["Пятница — День 2"]

    ensure_extra_press_block(day2)

    fill_entry(
        day1,
        "Жим в тренажере на грудь",
        "13.03.2026",
        1,
        [(15, 10, None), (25, 10, None), (35, 6, None)],
        "35x6",
        "Разминка: 5x15. Хотел 32.5, но поставил 35.",
    )
    fill_entry(
        day1,
        "Тяга chest-supported row",
        "13.03.2026",
        1,
        [(35, 10, None), (35, 12, None), (40, 10, None)],
        "40x10",
        "Разминка: 10x15.",
    )
    fill_entry(
        day1,
        "Разгибание ног",
        "13.03.2026",
        1,
        [(30, 12, None), (35, 10, None), (35, 9, None)],
        "35x10",
        "Разминка: 25x8.",
    )
    fill_entry(
        day1,
        "Сгибание ног",
        "13.03.2026",
        1,
        [(45, 9, None), (50, 10, None), (55, 8, None)],
        "55x8",
        "",
    )
    fill_entry(
        day1,
        "Махи в стороны",
        "13.03.2026",
        1,
        [(8, 10, None), (8, 9, None)],
        "8x10",
        "",
    )
    fill_entry(
        day1,
        "Rear delt / обратные разведения",
        "13.03.2026",
        1,
        [(20, 12, None), (22.5, 12, None), (27.5, 10, 2)],
        "Фейс-пул 27.5x10",
        "Упражнение делал как фейс-пул. 27.5 было зря: осталось только 2 повтора в запасе.",
    )
    fill_entry(
        day1,
        "Третий блок плеч",
        "13.03.2026",
        1,
        [(20, 10, None), (15, 14, None), (15, 14, None)],
        "Тяга к подбородку 20x10",
        "Делал тягу к подбородку прямым грифом. После упражнения заболела спина.",
    )
    fill_entry(
        day1,
        "Бицепс",
        "13.03.2026",
        1,
        [(20, 13, None), (20, 10, None), (20, 10, None)],
        "20x13",
        "Прямая штанга.",
    )
    fill_entry(
        day1,
        "Пресс",
        "13.03.2026",
        1,
        [(27.5, 18, None), (30, 20, None), (35, 20, 0)],
        "35x20",
        "Последний подход почти до отказа.",
    )

    fill_entry(
        day2,
        "Тяга вертикального блока",
        "17.03.2026",
        1,
        [(50, 9, None), (50, 9, None), (55, 7, None)],
        "55x7",
        "Разминка: 30x15.",
    )
    fill_entry(
        day2,
        "Жим в тренажере на грудь",
        "17.03.2026",
        1,
        [(20, 10, None), (27.5, 9, None), (27.5, 7, 2)],
        "27.5x9",
        "Разминка: 5x15. На последнем подходе ощущалось около 2 повторов в запасе.",
    )
    fill_entry(
        day2,
        "Жим ногами",
        "17.03.2026",
        1,
        [(100, 12, None), (130, 12, None), (150, 10, None), (170, 9, None)],
        "170x9",
        "Разминка: 40x15. Доп. подводящий сет: 80x12. На 170 казалось, что халтурил.",
    )
    fill_entry(
        day2,
        "Гиперэкстензия",
        "17.03.2026",
        1,
        [(0, 15, None)],
        "15 без веса",
        "",
    )
    fill_entry(
        day2,
        "Махи в стороны",
        "17.03.2026",
        1,
        [(7, 12, None), (7, 15, None), (8, 10, None), (8, 11, None)],
        "8x11",
        "",
    )
    fill_entry(
        day2,
        "Трицепс",
        "17.03.2026",
        1,
        [(20, 12, None), (22.5, 6, 1), (17.5, 8, None)],
        "20x12",
        "Канат. 22.5 шло тяжело, в запасе осталось 1-2 повтора.",
    )
    fill_entry(
        day2,
        "Бицепс",
        "17.03.2026",
        1,
        [(20, 14, None), (20, 9, None), (20, 8, None)],
        "20x14",
        "Прямая штанга.",
    )
    fill_entry(
        day2,
        "Пресс (дополнительно)",
        "17.03.2026",
        1,
        [],
        "До отказа",
        "Пресс выполнялся дополнительно, без точных веса и повторов.",
    )

    wb.save(WORKBOOK_PATH)
    wb.close()


if __name__ == "__main__":
    main()
