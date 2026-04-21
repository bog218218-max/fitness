#!/Users/bogdanromanenko/Desktop/тренировки/.venv/bin/python
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


WORKBOOK_PATH = Path("/Users/bogdanromanenko/Desktop/тренировки/программа.xlsx")

HEADERS = [
    "Дата",
    "Неделя",
    "Вес 1",
    "Повт 1",
    "RIR 1",
    "Вес 2",
    "Повт 2",
    "RIR 2",
    "Вес 3",
    "Повт 3",
    "RIR 3",
    "Вес 4",
    "Повт 4",
    "RIR 4",
    "Лучший сет / итог",
    "Заметки",
]

COLUMN_WIDTHS = {
    "A": 14,
    "B": 10,
    "C": 10,
    "D": 9,
    "E": 8,
    "F": 10,
    "G": 9,
    "H": 8,
    "I": 10,
    "J": 9,
    "K": 8,
    "L": 10,
    "M": 9,
    "N": 8,
    "O": 18,
    "P": 32,
}


def rir_sequence(load_type: str) -> list[int]:
    mapping = {
        "Тяжелая": [3, 2, 1, 3],
        "Легкая": [4, 3, 2, 4],
        "Средняя": [3, 2, 1, 3],
    }
    return mapping[load_type]


TUESDAY_PLAN = [
    {
        "title": "Жим в тренажере на грудь | 3x6-10, тяжелая",
        "type": "Тяжелая",
        "sets": [(30, 10), (30, 9), (30, 8)],
        "summary": "30x10",
        "notes": "",
    },
    {
        "title": "Тяга chest-supported row | 3x10-15, легкая",
        "type": "Легкая",
        "sets": [(37.5, 12), (37.5, 11), (37.5, 10)],
        "summary": "37.5x12",
        "notes": "Если нет 37.5: 35 x 13 / 12 / 11",
    },
    {
        "title": "Разгибание ног | 2-3x10-15, средняя",
        "type": "Средняя",
        "sets": [(30, 15), (30, 13), (30, 12)],
        "summary": "30x15",
        "notes": "",
    },
    {
        "title": "Сгибание ног | 2-3x10-15, средняя",
        "type": "Средняя",
        "sets": [(45, 12), (45, 11), (45, 10)],
        "summary": "45x12",
        "notes": "",
    },
    {
        "title": "Махи в стороны | 3x12-15, средняя",
        "type": "Средняя",
        "sets": [(7, 15), (7, 13), (7, 12)],
        "summary": "7x15",
        "notes": "",
    },
    {
        "title": "Фейс пул | 3x12-15, средняя",
        "type": "Средняя",
        "sets": [(20, 15), (20, 14), (20, 12)],
        "summary": "20x15",
        "notes": "",
    },
    {
        "title": "Тяга к подбородку | 3x15-20, средняя",
        "type": "Средняя",
        "sets": [(10, 15), (10, 13), (10, 12)],
        "summary": "10x15",
        "notes": "Хват чуть шире плеч. Без раскачки. Не тяни слишком высоко. Если снова чувствуется спина или неприятно плечу — убираем.",
    },
    {
        "title": "Бицепс | 3x8-12, средняя",
        "type": "Средняя",
        "sets": [(20, 12), (20, 10), (20, 9)],
        "summary": "20x12",
        "notes": "Прямая штанга.",
    },
    {
        "title": "Пресс | 2-3x15-20, средняя",
        "type": "Средняя",
        "sets": [(30, 18), (30, 16), (30, 15)],
        "summary": "30x18",
        "notes": "Не в отказ.",
    },
]

FRIDAY_PLAN = [
    {
        "title": "Тяга вертикального блока | 3x6-10, тяжелая",
        "type": "Тяжелая",
        "sets": [(50, 10), (50, 9), (50, 8)],
        "summary": "50x10",
        "notes": "",
    },
    {
        "title": "Жим в тренажере на грудь | 3x10-15, легкая",
        "type": "Легкая",
        "sets": [(22.5, 12), (22.5, 11), (22.5, 10)],
        "summary": "22.5x12",
        "notes": "",
    },
    {
        "title": "Жим ногами | 3x8-12, средняя",
        "type": "Средняя",
        "sets": [(150, 10), (150, 10), (150, 9)],
        "summary": "150x10",
        "notes": "",
    },
    {
        "title": "Гиперэкстензия | 2-3x12-15, средняя",
        "type": "Средняя",
        "sets": [(0, 15), (0, 15), (0, 15)],
        "summary": "15 без веса",
        "notes": "Свой вес.",
    },
    {
        "title": "Махи в стороны | 4x12-20, средняя",
        "type": "Средняя",
        "sets": [(7, 15), (7, 15), (7, 13), (7, 12)],
        "summary": "7x15",
        "notes": "",
    },
    {
        "title": "Трицепс | 3x10-15, средняя",
        "type": "Средняя",
        "sets": [(15, 15), (15, 12), (15, 10)],
        "summary": "15x15",
        "notes": "Канат.",
    },
    {
        "title": "Бицепс | 2x10-15, легкая",
        "type": "Легкая",
        "sets": [(17.5, 12), (17.5, 10)],
        "summary": "17.5x12",
        "notes": "Прямая штанга. Если хочешь 3 подхода: 17.5 x 12 / 10 / 9.",
    },
    {
        "title": "Пресс | 3x15, средняя",
        "type": "Средняя",
        "sets": [(30, 15), (30, 15), (30, 15)],
        "summary": "30x15",
        "notes": "",
    },
]


def apply_base_layout(ws, title: str) -> None:
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "A5"

    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:P1")

    ws["A2"] = "Теория: только рабочие подходы. Таблица сделана в том же формате, что и практика."
    ws["A3"] = "Недели 1-4 уже предзаполнены как цель цикла. Дату можно не трогать."
    for cell in (ws["A2"], ws["A3"]):
        cell.font = Font(italic=True, color="1F2937")
        cell.alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:P2")
    ws.merge_cells("A3:P3")

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 20


def write_headers(ws, row: int) -> None:
    for idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row, idx, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 22


def write_section_title(ws, row: int, title: str) -> None:
    cell = ws.cell(row, 1, title)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="0F766E")
    cell.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=16)
    ws.row_dimensions[row].height = 20


def write_data_row(ws, row: int, values: dict[int, object]) -> None:
    for col in range(1, 17):
        cell = ws.cell(row, col, values.get(col, ""))
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 18


def week_row_values(week: int, load_type: str, sets: list[tuple[float | int, int]], summary: str, notes: str) -> dict[int, object]:
    values: dict[int, object] = {
        2: week,
    }

    rir = rir_sequence(load_type)[week - 1]
    target_sets = sets if week == 1 else []

    col = 3
    for weight, reps in target_sets[:4]:
        values[col] = weight
        values[col + 1] = reps
        values[col + 2] = rir
        col += 3

    values[15] = summary if week == 1 else ""
    values[16] = notes if week == 1 else ""

    if week != 1:
        # Keep only the weekly target RIR visible when exact numbers are not yet assigned.
        for rir_col in (5, 8, 11, 14):
            values[rir_col] = rir

    return values


def placeholder_row_values() -> dict[int, object]:
    return {
        2: "—",
        3: "—",
        4: "—",
        5: "—",
        6: "—",
        7: "—",
        8: "—",
        9: "—",
        10: "—",
        11: "—",
        12: "—",
        13: "—",
        14: "—",
        15: "—",
        16: "Плана заранее не было.",
    }


def build_theory_sheet(ws, title: str, plan: list[dict]) -> None:
    apply_base_layout(ws, title)
    current_row = 5

    for exercise in plan:
        write_section_title(ws, current_row, exercise["title"])
        current_row += 1
        write_headers(ws, current_row)
        current_row += 1

        write_data_row(ws, current_row, placeholder_row_values())
        current_row += 1

        for week in range(1, 5):
            values = week_row_values(
                week=week,
                load_type=exercise["type"],
                sets=exercise["sets"],
                summary=exercise["summary"],
                notes=exercise["notes"],
            )
            write_data_row(ws, current_row, values)
            current_row += 1

        for _ in range(5):
            write_data_row(ws, current_row, {})
            current_row += 1

        current_row += 1


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH)

    if "Вторник — Практика" in wb.sheetnames:
        wb["Вторник — Практика"]["A1"] = "Вторник — Практика"
    if "Пятница — Практика" in wb.sheetnames:
        wb["Пятница — Практика"]["A1"] = "Пятница — Практика"

    for name in ["Вторник — Теория", "Пятница — Теория"]:
        if name in wb.sheetnames:
            del wb[name]

    ws_tue = wb.create_sheet("Вторник — Теория")
    build_theory_sheet(ws_tue, "Вторник — Теория", TUESDAY_PLAN)

    ws_fri = wb.create_sheet("Пятница — Теория")
    build_theory_sheet(ws_fri, "Пятница — Теория", FRIDAY_PLAN)

    wb.save(WORKBOOK_PATH)
    wb.close()

    verify = load_workbook(WORKBOOK_PATH, data_only=True)
    verify.close()


if __name__ == "__main__":
    main()
