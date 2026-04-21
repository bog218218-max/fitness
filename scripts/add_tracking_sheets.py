#!/Users/bogdanromanenko/Desktop/тренировки/.venv/bin/python
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


WORKBOOK_PATH = Path("/Users/bogdanromanenko/Desktop/тренировки/программа.xlsx")
BACKUP_PATH = Path("/Users/bogdanromanenko/Desktop/тренировки/программа.backup.xlsx")

TRACKING_LAYOUT = [
    (
        "Вторник — День 1",
        [
            ("Жим в тренажере на грудь", "3x6-10, тяжелая"),
            ("Тяга chest-supported row", "3x10-15, легкая"),
            ("Разгибание ног", "2-3x10-15, средняя"),
            ("Сгибание ног", "2-3x10-15, средняя"),
            ("Махи в стороны", "3x12-15, средняя"),
            ("Rear delt / обратные разведения", "3x12-15, средняя"),
            ("Третий блок плеч", "3x15-20, средняя"),
            ("Бицепс", "3x8-12, средняя"),
            ("Пресс", "2-3x15-20, средняя"),
        ],
    ),
    (
        "Пятница — День 2",
        [
            ("Тяга вертикального блока", "3x6-10, тяжелая"),
            ("Жим в тренажере на грудь", "3x10-15, легкая"),
            ("Жим ногами", "3x8-12, средняя"),
            ("Гиперэкстензия", "2-3x12-15, средняя"),
            ("Махи в стороны", "4x12-20, средняя"),
            ("Трицепс", "3x10-15, средняя"),
            ("Бицепс", "2x10-15, легкая"),
        ],
    ),
]

HEADERS = [
    "Дата",
    "Неделя",
    "Вес 1",
    "Повт 1",
    "RIR 1",
    "Вес 2",
    "Повт 2",
    "RIR 2",
    "Вес 3",
    "Повт 3",
    "RIR 3",
    "Вес 4",
    "Повт 4",
    "RIR 4",
    "Лучший сет / итог",
    "Заметки",
]

COLUMN_WIDTHS = {
    "A": 14,
    "B": 10,
    "C": 10,
    "D": 9,
    "E": 8,
    "F": 10,
    "G": 9,
    "H": 8,
    "I": 10,
    "J": 9,
    "K": 8,
    "L": 10,
    "M": 9,
    "N": 8,
    "O": 18,
    "P": 32,
}


def apply_layout(ws, title: str, exercises: list[tuple[str, str]]) -> None:
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "A5"

    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:P1")

    ws["A2"] = "Заполняй только рабочие подходы. Разминку лучше не смешивать с рабочими сетами."
    ws["A3"] = "RIR = сколько повторов осталось в запасе в конце подхода."
    for cell in (ws["A2"], ws["A3"]):
        cell.alignment = Alignment(wrap_text=True)
        cell.font = Font(italic=True, color="1F2937")
    ws.merge_cells("A2:P2")
    ws.merge_cells("A3:P3")

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 20

    current_row = 5
    for exercise, target in exercises:
        title_cell = ws.cell(current_row, 1, f"{exercise} | {target}")
        title_cell.font = Font(bold=True, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor="0F766E")
        title_cell.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=16)
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        for idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(current_row, idx, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        for _ in range(10):
            for idx in range(1, 17):
                ws.cell(current_row, idx, "")
            ws.row_dimensions[current_row].height = 18
            current_row += 1

        current_row += 1


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(WORKBOOK_PATH)
    if not BACKUP_PATH.exists():
        raise FileNotFoundError(BACKUP_PATH)

    # Start from the clean backup to avoid carrying over the corrupted workbook state.
    WORKBOOK_PATH.write_bytes(BACKUP_PATH.read_bytes())

    wb = load_workbook(WORKBOOK_PATH)

    for name in [sheet_name for sheet_name, _ in TRACKING_LAYOUT]:
        if name in wb.sheetnames:
            del wb[name]

    for sheet_name, exercises in TRACKING_LAYOUT:
        ws = wb.create_sheet(title=sheet_name)
        apply_layout(ws, sheet_name, exercises)

    wb.save(WORKBOOK_PATH)

    # Re-open once to ensure the file is readable after save.
    verify = load_workbook(WORKBOOK_PATH, data_only=True)
    verify.close()


if __name__ == "__main__":
    main()
