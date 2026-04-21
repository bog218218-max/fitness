#!/Users/bogdanromanenko/Desktop/тренировки/.venv/bin/python
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path("/Users/bogdanromanenko/Desktop/тренировки/программа.xlsx")


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH)

    rename_map = {
        "Вторник — День 1": "Вторник — Практика",
        "Пятница — День 2": "Пятница — Практика",
    }

    for old_name, new_name in rename_map.items():
        if old_name in wb.sheetnames:
            wb[old_name].title = new_name

    wb.save(WORKBOOK_PATH)
    wb.close()

    verify = load_workbook(WORKBOOK_PATH, data_only=True)
    verify.close()


if __name__ == "__main__":
    main()
