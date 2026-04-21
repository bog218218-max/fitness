#!/Users/bogdanromanenko/Desktop/тренировки/.venv/bin/python
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path("/Users/bogdanromanenko/Desktop/тренировки/программа.xlsx")


def find_row(ws, title_prefix: str) -> int | None:
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if isinstance(value, str) and value.startswith(title_prefix):
            return row
    return None


def clear_section(ws, title_row: int) -> None:
    next_section = ws.max_row + 1
    for row in range(title_row + 1, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if row > title_row and isinstance(value, str) and "|" in value:
            next_section = row
            break

    merged_ranges = list(ws.merged_cells.ranges)
    for merged in merged_ranges:
        if merged.min_row >= title_row and merged.max_row < next_section:
            ws.unmerge_cells(str(merged))

    for row in range(title_row, next_section):
        for col in range(1, 17):
            ws.cell(row, col).value = None


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH)

    for sheet_name in ["Вторник — Отчет"]:
        ws = wb[sheet_name]

        rear_delt_row = find_row(ws, "Rear delt / обратные разведения")
        if rear_delt_row:
            ws.cell(rear_delt_row, 1).value = "Махи задняя дельта на скамье | 3x12-15, средняя"

        third_block_row = find_row(ws, "Третий блок плеч")
        if third_block_row:
            clear_section(ws, third_block_row)

    wb.save(WORKBOOK_PATH)
    wb.close()


if __name__ == "__main__":
    main()
