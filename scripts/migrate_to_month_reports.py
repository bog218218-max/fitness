#!/Users/bogdanromanenko/Desktop/тренировки/.venv/bin/python
from __future__ import annotations

from pathlib import Path
from shutil import copyfile

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


WORKBOOK_PATH = Path("/Users/bogdanromanenko/Desktop/тренировки/программа.xlsx")
BACKUP_PATH = Path("/Users/bogdanromanenko/Desktop/тренировки/программа.monthly-report.backup.xlsx")

HEADERS = [
    "Дата",
    "Месяц",
    "Вес 1",
    "Повт 1",
    "RIR 1",
    "Вес 2",
    "Повт 2",
    "RIR 2",
    "Вес 3",
    "Повт 3",
    "RIR 3",
    "Вес 4",
    "Повт 4",
    "RIR 4",
    "Итог",
    "Заметки",
]

TUESDAY_MONTH_1 = {
    "Жим в тренажере на грудь": {
        "sets": [(30, 10), (35, 9), (37.5, 6)],
        "summary": "37.5x6",
        "notes": "",
    },
    "Тяга chest-supported row": {
        "sets": [(40, 13), (45, 11), (47.5, 10)],
        "summary": "47.5x10",
        "notes": "",
    },
    "Разгибание ног": {
        "sets": [(35, 10), (40, 13), (45, 12)],
        "summary": "45x12",
        "notes": "",
    },
    "Сгибание ног": {
        "sets": [(50, 10), (55, 8), (50, 12)],
        "summary": "55x8",
        "notes": "",
    },
    "Махи в стороны": {
        "sets": [(7, 14), (8, 12), (10, 8)],
        "summary": "10x8",
        "notes": "",
    },
    "Rear delt / обратные разведения": {
        "sets": [(7, 12), (10, 8), (8, 12)],
        "summary": "10x8",
        "notes": "Задняя дельта на скамье.",
    },
    "Третий блок плеч": {
        "sets": [],
        "summary": "",
        "notes": "Нет записи в отчете за месяц 1.",
    },
    "Бицепс": {
        "sets": [(20, 12), (25, 10)],
        "summary": "25x10",
        "notes": "",
    },
    "Пресс": {
        "sets": [],
        "summary": "",
        "notes": "Нет записи в отчете за месяц 1.",
    },
}

FRIDAY_MONTH_1 = {
    "Тяга вертикального блока": {
        "sets": [(60, 10), (65, 8), (70, 8)],
        "summary": "70x8",
        "notes": "",
    },
    "Жим в тренажере на грудь": {
        "sets": [(25, 13), (30, 10), (30, 8)],
        "summary": "30x10",
        "notes": "",
    },
    "Жим ногами": {
        "sets": [(80, 15), (100, 12)],
        "summary": "100x12",
        "notes": "",
    },
    "Гиперэкстензия": {
        "sets": [],
        "summary": "",
        "notes": "Нет записи в отчете за месяц 1.",
    },
    "Махи в стороны": {
        "sets": [(7, 18), (8, 15), (9, 15)],
        "summary": "9x15",
        "notes": "",
    },
    "Трицепс": {
        "sets": [(17.5, 12), (17.5, 12), (20, 11)],
        "summary": "20x11",
        "notes": "",
    },
    "Бицепс": {
        "sets": [(25, 10), (20, 12)],
        "summary": "25x10",
        "notes": "",
    },
    "Пресс": {
        "sets": [],
        "summary": "",
        "notes": "Нет записи в отчете за месяц 1.",
    },
}


def normalize_title(value: str | None) -> str:
    return (value or "").split("|")[0].strip()


def section_rows(ws) -> list[int]:
    return [
        row
        for row in range(1, ws.max_row + 1)
        if isinstance(ws.cell(row, 1).value, str) and "|" in ws.cell(row, 1).value
    ]


def update_sheet_intro(ws, title: str) -> None:
    ws["A1"] = title
    ws["A2"] = "Отчет по месяцу: записываем только одну контрольную неделю, обычно 3-ю."
    ws["A3"] = "В таблице только факт. Целей и теоретических листов больше нет."
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(italic=True, color="1F2937")
    ws["A3"].font = Font(italic=True, color="1F2937")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws["A3"].alignment = Alignment(wrap_text=True)


def clear_data_block(ws, start_row: int, end_row: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(1, 17):
            ws.cell(row, col).value = None


def write_headers(ws, header_row: int) -> None:
    for index, header in enumerate(HEADERS, start=1):
        ws.cell(header_row, index, header)


def write_month_slot(ws, row: int, month: int) -> None:
    ws.cell(row, 1, "")
    ws.cell(row, 2, month)


def write_entry(ws, row: int, month: int, entry: dict[str, object]) -> None:
    ws.cell(row, 2, month)
    sets = entry["sets"]
    col = 3
    for weight, reps in sets[:4]:
        ws.cell(row, col, weight)
        ws.cell(row, col + 1, reps)
        col += 3
    ws.cell(row, 15, entry["summary"])
    ws.cell(row, 16, entry["notes"])


def migrate_report_sheet(ws, title: str, month_1_data: dict[str, dict[str, object]]) -> None:
    update_sheet_intro(ws, title)

    for section_row in section_rows(ws):
        header_row = section_row + 1
        data_start = section_row + 2
        data_end = section_row + 11
        exercise_name = normalize_title(ws.cell(section_row, 1).value)

        if exercise_name == "Пресс (дополнительно)":
            ws.cell(section_row, 1, "Пресс | 2-3x15-20, средняя")
            exercise_name = "Пресс"

        write_headers(ws, header_row)
        clear_data_block(ws, data_start, data_end)

        for offset, month in enumerate(range(1, 7)):
            write_month_slot(ws, data_start + offset, month)

        entry = month_1_data.get(exercise_name)
        if entry:
            write_entry(ws, data_start, 1, entry)


def main() -> None:
    copyfile(WORKBOOK_PATH, BACKUP_PATH)
    wb = load_workbook(WORKBOOK_PATH)

    for old_name, new_name in {
        "Вторник — Практика": "Вторник — Отчет",
        "Пятница — Практика": "Пятница — Отчет",
    }.items():
        if old_name in wb.sheetnames:
            wb[old_name].title = new_name

    for name in ["Вторник — Теория", "Пятница — Теория"]:
        if name in wb.sheetnames:
            del wb[name]

    if "Вторник — Отчет" in wb.sheetnames:
        migrate_report_sheet(wb["Вторник — Отчет"], "Вторник — Отчет", TUESDAY_MONTH_1)
    if "Пятница — Отчет" in wb.sheetnames:
        migrate_report_sheet(wb["Пятница — Отчет"], "Пятница — Отчет", FRIDAY_MONTH_1)

    wb.save(WORKBOOK_PATH)
    wb.close()


if __name__ == "__main__":
    main()
