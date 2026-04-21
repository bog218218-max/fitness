#!/Users/bogdanromanenko/Desktop/тренировки/.venv/bin/python
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path("/Users/bogdanromanenko/Desktop/тренировки/программа.xlsx")


def find_title_row(ws, title_prefix: str) -> int:
    for row in range(1, ws.max_row + 1):
        value = ws[f"A{row}"].value
        if isinstance(value, str) and value.startswith(title_prefix):
            return row
    raise ValueError(f"Title not found: {title_prefix}")


def fill_default_rir(ws, title_prefix: str, default_rir: int) -> None:
    row = find_title_row(ws, title_prefix) + 2
    for rir_col, weight_col in ((5, 3), (8, 6), (11, 9), (14, 12)):
        weight = ws.cell(row, weight_col).value
        if weight is not None and ws.cell(row, rir_col).value is None:
            ws.cell(row, rir_col, default_rir)


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH)

    day1 = wb["Вторник — День 1"]
    day2 = wb["Пятница — День 2"]

    # Week 1 defaults from the plan.
    for title in [
        "Жим в тренажере на грудь",
    ]:
        fill_default_rir(day1, title, 3)

    for title in [
        "Тяга chest-supported row",
    ]:
        fill_default_rir(day1, title, 4)

    for title in [
        "Разгибание ног",
        "Сгибание ног",
        "Махи в стороны",
        "Rear delt / обратные разведения",
        "Третий блок плеч",
        "Бицепс",
        "Пресс",
    ]:
        fill_default_rir(day1, title, 3)

    for title in [
        "Тяга вертикального блока",
    ]:
        fill_default_rir(day2, title, 3)

    for title in [
        "Жим в тренажере на грудь",
        "Бицепс",
    ]:
        fill_default_rir(day2, title, 4)

    for title in [
        "Жим ногами",
        "Гиперэкстензия",
        "Махи в стороны",
        "Трицепс",
    ]:
        fill_default_rir(day2, title, 3)

    wb.save(WORKBOOK_PATH)
    wb.close()


if __name__ == "__main__":
    main()
