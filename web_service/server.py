#!/Users/bogdanromanenko/Desktop/тренировки/.venv/bin/python
from __future__ import annotations

import hashlib
import hmac
import json
import os
import re
import secrets
import shutil
import time
from dataclasses import asdict, dataclass
from datetime import datetime
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from threading import Lock
from typing import Any
from urllib.parse import urlparse

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
STATIC_DIR = ROOT / "static"
WORKBOOK_PATH = ROOT.parent / "программа.xlsx"
GENERAL_BACKUP_PATH = ROOT.parent / "программа.backup.xlsx"
MONTHLY_REPORT_BACKUP_PATH = ROOT.parent / "программа.monthly-report.backup.xlsx"
HOST = os.environ.get("FITNESS_HOST", "0.0.0.0")
PORT = int(os.environ.get("FITNESS_PORT", "8123"))
ADMIN_PIN = os.environ.get("FITNESS_ADMIN_PIN", "1111")
TOKEN_SECRET = os.environ.get("FITNESS_ADMIN_TOKEN_SECRET", secrets.token_hex(32)).encode("utf-8")
TOKEN_TTL_SECONDS = 60 * 60 * 24 * 30
WORKBOOK_LOCK = Lock()


@dataclass
class SetData:
    weight: float | int | None
    reps: int | None


@dataclass
class EntryData:
    date: str | None
    month: int | str | None
    summary: str | None
    notes: str | None
    sets: list[SetData]
    logged: bool
    best_set: str | None
    estimated_1rm: float | None


class RequestError(Exception):
    def __init__(self, message: str, status: HTTPStatus = HTTPStatus.BAD_REQUEST):
        super().__init__(message)
        self.message = message
        self.status = status


def normalize_number(value: Any) -> float | int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else float(value)
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped or stripped == "—":
            return None
        try:
            number = float(stripped.replace(",", "."))
        except ValueError:
            return None
        return int(number) if number.is_integer() else number
    return None


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return str(value)


def normalize_key(value: Any) -> str:
    text = normalize_text(value) or ""
    return re.sub(r"\s+", " ", text.replace("ё", "е").replace("Ё", "Е").strip().lower())


def is_section_title(value: Any) -> bool:
    return isinstance(value, str) and "|" in value and not value.startswith("Дата")


def format_number(value: float | int | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, int):
        return str(value)
    return str(int(value)) if float(value).is_integer() else str(value).rstrip("0").rstrip(".")


def estimate_1rm(weight: float | int | None, reps: int | None) -> float | None:
    if weight is None or reps is None or weight <= 0 or reps <= 0:
        return None
    return float(weight) * (1 + int(reps) / 30)


def build_best_set(sets: list[SetData]) -> tuple[str | None, float | None]:
    best_set = None
    best_1rm = None

    for set_data in sets:
        current_1rm = estimate_1rm(set_data.weight, set_data.reps)
        if current_1rm is None:
            continue
        if best_1rm is None or current_1rm > best_1rm:
            best_1rm = current_1rm
            best_set = f"{format_number(set_data.weight)}x{set_data.reps}"

    return best_set, round(best_1rm, 1) if best_1rm is not None else None


def parse_exercise_title(raw_title: str | None) -> dict[str, str | None]:
    title = normalize_text(raw_title)
    if not title:
        return {"base_title": None, "prescription": None, "load_type": None}

    base_title, _, tail = title.partition("|")
    prescription = normalize_text(tail)
    load_type = None
    if prescription and "," in prescription:
        _, _, load_fragment = prescription.rpartition(",")
        load_type = normalize_text(load_fragment)

    return {
        "base_title": normalize_text(base_title),
        "prescription": prescription,
        "load_type": load_type,
    }


def parse_rir_label(value: Any) -> str | None:
    text = normalize_text(value)
    if not text:
        return None
    match = re.search(r"(\d+(?:\.\d+)?)", text)
    return match.group(1) if match else text


def build_rir_lookup(sheet) -> dict[str, str]:
    rir_lookup: dict[str, str] = {}
    for row_index in range(1, sheet.max_row + 1):
        load_type = normalize_text(sheet.cell(row_index, 1).value)
        if not load_type:
            continue

        weeks = [parse_rir_label(sheet.cell(row_index, column).value) for column in range(2, 6)]
        weeks = [week for week in weeks if week]
        if not weeks:
            continue

        rir_lookup[normalize_key(load_type)] = "→".join(weeks)
    return rir_lookup


def resolve_target_rir(exercise_title: str | None, rir_lookup: dict[str, str]) -> tuple[str | None, str | None, str | None]:
    parsed = parse_exercise_title(exercise_title)
    load_type = parsed["load_type"]
    target_rir = rir_lookup.get(normalize_key(load_type)) if load_type else None
    return parsed["base_title"], parsed["prescription"], target_rir


def parse_entry(row: list[Any]) -> EntryData:
    month_value = row[1]
    month = month_value if isinstance(month_value, str) else normalize_number(month_value)
    date = normalize_text(row[0])
    summary = normalize_text(row[14])
    notes = normalize_text(row[15])

    sets: list[SetData] = []
    for start in (2, 5, 8, 11):
        weight = normalize_number(row[start])
        reps = normalize_number(row[start + 1])
        sets.append(SetData(weight=weight, reps=int(reps) if reps is not None else None))

    best_set, best_1rm = build_best_set(sets)
    logged = bool(summary or notes or any(item.weight is not None or item.reps is not None for item in sets))

    return EntryData(
        date=date,
        month=month,
        summary=summary,
        notes=notes,
        sets=sets,
        logged=logged,
        best_set=best_set,
        estimated_1rm=best_1rm,
    )


def parse_structured_sheet(sheet, rir_lookup: dict[str, str]) -> dict[str, Any]:
    section_rows = [
        row_index
        for row_index in range(1, sheet.max_row + 1)
        if is_section_title(sheet.cell(row_index, 1).value)
    ]

    exercises = []
    for index, section_row in enumerate(section_rows):
        next_section = section_rows[index + 1] if index + 1 < len(section_rows) else sheet.max_row + 1
        title = normalize_text(sheet.cell(section_row, 1).value)
        base_title, prescription, target_rir = resolve_target_rir(title, rir_lookup)
        load_type = parse_exercise_title(title)["load_type"]
        entries: list[EntryData] = []
        for row_index in range(section_row + 2, next_section):
            row = [sheet.cell(row_index, column).value for column in range(1, 17)]
            if all(value in (None, "") for value in row):
                continue
            if row[0] == "Дата":
                continue
            entries.append(parse_entry(row))
        exercises.append(
            {
                "title": title,
                "base_title": base_title,
                "prescription": prescription,
                "load_type": load_type,
                "target_rir": target_rir,
                "entries": [asdict(entry) for entry in entries],
            }
        )
    return {"title": sheet.title, "exercises": exercises}


def load_dashboard() -> dict[str, Any]:
    with WORKBOOK_LOCK:
        workbook = load_workbook(WORKBOOK_PATH, data_only=True)
        rir_lookup = build_rir_lookup(workbook["RIR и прогрессия"]) if "RIR и прогрессия" in workbook.sheetnames else {}
        reports = {}
        for sheet_name in workbook.sheetnames:
            if "Отчет" in sheet_name:
                reports[sheet_name] = parse_structured_sheet(workbook[sheet_name], rir_lookup)
        workbook.close()

    stat = WORKBOOK_PATH.stat()
    return {
        "workbook": {
            "path": str(WORKBOOK_PATH),
            "updated_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            "size_bytes": stat.st_size,
            "backups": {
                "general": str(GENERAL_BACKUP_PATH),
                "monthly_report": str(MONTHLY_REPORT_BACKUP_PATH),
            },
        },
        "reports": reports,
    }


def read_json_body(handler: SimpleHTTPRequestHandler) -> dict[str, Any]:
    length = int(handler.headers.get("Content-Length", "0"))
    raw_body = handler.rfile.read(length) if length else b"{}"
    try:
        payload = json.loads(raw_body.decode("utf-8") or "{}")
    except json.JSONDecodeError as exc:
        raise RequestError("Некорректный JSON в запросе.") from exc
    if not isinstance(payload, dict):
        raise RequestError("Тело запроса должно быть JSON-объектом.")
    return payload


def issue_admin_token() -> dict[str, str]:
    expires_at = int(time.time()) + TOKEN_TTL_SECONDS
    payload = str(expires_at)
    signature = hmac.new(TOKEN_SECRET, payload.encode("utf-8"), hashlib.sha256).hexdigest()
    return {"token": f"{payload}.{signature}", "expires_at": datetime.fromtimestamp(expires_at).isoformat()}


def validate_admin_token(token: str | None) -> bool:
    if not token or "." not in token:
        return False
    expires_raw, provided_signature = token.split(".", 1)
    if not expires_raw.isdigit():
        return False
    if int(expires_raw) < int(time.time()):
        return False

    expected_signature = hmac.new(TOKEN_SECRET, expires_raw.encode("utf-8"), hashlib.sha256).hexdigest()
    return hmac.compare_digest(expected_signature, provided_signature)


def get_bearer_token(handler: SimpleHTTPRequestHandler) -> str | None:
    authorization = handler.headers.get("Authorization", "")
    prefix = "Bearer "
    if authorization.startswith(prefix):
        return authorization[len(prefix) :].strip()
    return None


def require_admin(handler: SimpleHTTPRequestHandler) -> None:
    if not validate_admin_token(get_bearer_token(handler)):
        raise RequestError("Нужен вход в админку.", status=HTTPStatus.UNAUTHORIZED)


def sanitize_sets(raw_sets: Any) -> list[SetData]:
    if raw_sets is None:
        return []
    if not isinstance(raw_sets, list):
        raise RequestError("Подходы должны приходить списком.")
    if len(raw_sets) > 4:
        raise RequestError("Можно сохранить не больше 4 рабочих подходов.")

    normalized_sets: list[SetData] = []
    for item in raw_sets:
        if not isinstance(item, dict):
            raise RequestError("Каждый подход должен быть объектом.")
        weight = normalize_number(item.get("weight"))
        reps = normalize_number(item.get("reps"))
        if weight is None and reps is None:
            normalized_sets.append(SetData(weight=None, reps=None))
            continue
        if weight is None or reps is None:
            raise RequestError("У подхода должны быть одновременно и вес, и повторы.")
        if weight <= 0 or reps <= 0:
            raise RequestError("Вес и повторы должны быть больше нуля.")
        normalized_sets.append(SetData(weight=weight, reps=int(reps)))

    while len(normalized_sets) < 4:
        normalized_sets.append(SetData(weight=None, reps=None))
    return normalized_sets


def find_entry_row(sheet, exercise_title: str, month: int) -> int | None:
    section_row = None
    for row_index in range(1, sheet.max_row + 1):
        if sheet.cell(row_index, 1).value == exercise_title:
            section_row = row_index
            break
    if section_row is None:
        return None

    row_index = section_row + 2
    while row_index <= sheet.max_row:
        title_cell = sheet.cell(row_index, 1).value
        if is_section_title(title_cell):
            break
        current_month = normalize_number(sheet.cell(row_index, 2).value)
        if current_month == month:
            return row_index
        row_index += 1
    return None


def ensure_backups() -> None:
    if not GENERAL_BACKUP_PATH.exists():
        shutil.copy2(WORKBOOK_PATH, GENERAL_BACKUP_PATH)
    shutil.copy2(WORKBOOK_PATH, MONTHLY_REPORT_BACKUP_PATH)


def save_report_entry(payload: dict[str, Any]) -> dict[str, Any]:
    sheet_name = normalize_text(payload.get("sheet_name"))
    exercise_title = normalize_text(payload.get("exercise_title"))
    month = normalize_number(payload.get("month"))
    notes = normalize_text(payload.get("notes"))
    sets = sanitize_sets(payload.get("sets"))

    if not sheet_name:
        raise RequestError("Не выбран день отчета.")
    if not exercise_title:
        raise RequestError("Не выбрано упражнение.")
    if month is None:
        raise RequestError("Не выбран месяц.")

    month_number = int(month)

    with WORKBOOK_LOCK:
        workbook = load_workbook(WORKBOOK_PATH)
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise RequestError("Лист отчета не найден.")
        sheet = workbook[sheet_name]
        target_row = find_entry_row(sheet, exercise_title, month_number)
        if target_row is None:
            workbook.close()
            raise RequestError("Не удалось найти строку месяца для этого упражнения.")

        ensure_backups()

        for index, set_data in enumerate(sets):
            column = 3 + index * 3
            sheet.cell(target_row, column).value = set_data.weight
            sheet.cell(target_row, column + 1).value = set_data.reps
            sheet.cell(target_row, column + 2).value = None

        best_set, _ = build_best_set(sets)
        sheet.cell(target_row, 15).value = best_set
        sheet.cell(target_row, 16).value = notes

        workbook.save(WORKBOOK_PATH)
        workbook.close()

    return {"ok": True, "saved": {"sheet_name": sheet_name, "exercise_title": exercise_title, "month": month_number}}


class AppHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(STATIC_DIR), **kwargs)

    def end_headers(self) -> None:
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        super().end_headers()

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        try:
            if parsed.path == "/api/dashboard":
                self._send_json(load_dashboard())
                return
            if parsed.path == "/api/health":
                self._send_json({"ok": True})
                return
            if parsed.path == "/api/admin/session":
                require_admin(self)
                self._send_json({"ok": True})
                return
            if parsed.path == "/":
                self.path = "/index.html"
            return super().do_GET()
        except RequestError as exc:
            self._send_json({"ok": False, "error": exc.message}, status=exc.status)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        try:
            if parsed.path == "/api/admin/login":
                payload = read_json_body(self)
                submitted_pin = normalize_text(payload.get("pin"))
                if submitted_pin != ADMIN_PIN:
                    raise RequestError("Неверный PIN.", status=HTTPStatus.UNAUTHORIZED)
                self._send_json({"ok": True, **issue_admin_token()})
                return

            if parsed.path == "/api/admin/save-entry":
                require_admin(self)
                payload = read_json_body(self)
                self._send_json(save_report_entry(payload))
                return

            self._send_json({"ok": False, "error": "Маршрут не найден."}, status=HTTPStatus.NOT_FOUND)
        except RequestError as exc:
            self._send_json({"ok": False, "error": exc.message}, status=exc.status)

    def _send_json(self, payload: dict[str, Any], status: HTTPStatus = HTTPStatus.OK) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


def main() -> None:
    server = ThreadingHTTPServer((HOST, PORT), AppHandler)
    print(f"Serving progress app at http://127.0.0.1:{PORT}")
    server.serve_forever()


if __name__ == "__main__":
    main()
